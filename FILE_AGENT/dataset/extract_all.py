from __future__ import annotations

import argparse
import math
import os
import re
import shutil
import sys
import tarfile
import time
import warnings
import zipfile
import gzip
from collections import Counter
from pathlib import Path

import pandas as pd

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# Đường dẫn mặc định
# ─────────────────────────────────────────────────────────────────────────────
SCRIPT_DIR  = Path(__file__).resolve().parent
DATASET_DIR = SCRIPT_DIR / "Dataset"

DEFAULT_OUT = {
    "pdf":   DATASET_DIR / "PDF"              / "PDF_Extract.csv",
    "word":  DATASET_DIR / "Word Document"    / "WORD_Extract.csv",
    "excel": DATASET_DIR / "Excel"            / "EXCEL_Extract.csv",
    "qr":    DATASET_DIR / "QR Codes"         / "QR_Extract.csv",
}

PDF_EXTS   = {".pdf"}
WORD_EXTS  = {".doc", ".docx", ".docm", ".dot", ".dotx", ".dotm"}
EXCEL_EXTS = {".xls", ".xlsx", ".xlsm", ".xlsb"}
QR_EXTS    = {".png", ".jpg", ".jpeg", ".bmp", ".gif", ".tiff", ".webp"}
ARCHIVE_EXTS = {".zip", ".tar", ".taz", ".tar.gz", ".tgz", ".gz"}

# ─────────────────────────────────────────────────────────────────────────────
# Shared utilities
# ─────────────────────────────────────────────────────────────────────────────

def _entropy(text: str) -> float:
    if not text:
        return 0.0
    probs = [n / len(text) for n in Counter(text).values()]
    return -sum(p * math.log2(p) for p in probs)


def _log(log_path: Path, filepath: str, msg: str) -> None:
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"{filepath} -- {msg}\n")


# ─────────────────────────────────────────────────────────────────────────────
# PDF extractor (40 features)
# ─────────────────────────────────────────────────────────────────────────────

PDF_COLS = [
    "file_path", "file_size", "title_chars", "encrypted", "metadata_size",
    "page_count", "valid_pdf_header", "image_count", "text_length",
    "object_count", "font_object_count", "embedded_file_count",
    "average_embedded_file_size", "stream_count", "endstream_count",
    "average_stream_size", "entropy_of_streams", "xref_count", "xref_entries",
    "name_obfuscations", "total_filters", "nested_filter_objects",
    "objstm_count", "js_count", "javascript_count", "uri_count",
    "uses_nonstandard_port", "action_count", "aa_count", "openaction_count",
    "launch_count", "submitform_count", "acroform_count", "xfa_count",
    "jbig2decode_count", "colors_count", "richmedia_count",
    "trailer_count", "startxref_count",
    "has_multiple_behavioral_keywords_in_one_object",
    "used_ocr", "label",
]

def _count_name_obfuscations(text: str) -> int:
    patterns = [
        r"/[a-zA-Z]*#\d{2}",
        r"/[a-zA-Z]*%[0-9a-fA-F]{2}",
        r"/[a-zA-Z]*\\x[0-9a-fA-F]{2}",
        r"/[a-zA-Z]*\\[0-7]{1,3}",
    ]
    return sum(len(re.findall(p, text)) for p in patterns)


def extract_pdf(filepath: str, label: int | None, log_path: Path) -> dict:
    feats = {col: 0 for col in PDF_COLS}
    feats["file_path"] = filepath
    feats["label"] = label if label is not None else ""
    try:
        feats["file_size"] = os.path.getsize(filepath)
    except OSError as e:
        _log(log_path, filepath, f"Size check failed: {e}")
        return feats

    # PyPDF2 stage
    try:
        import PyPDF2
        with open(filepath, "rb") as f:
            reader = PyPDF2.PdfReader(f, strict=False)
            feats["encrypted"] = int(reader.is_encrypted)
            feats["page_count"] = len(reader.pages)
            meta = reader.metadata
            if meta:
                feats["metadata_size"] = len(str(meta))
                feats["title_chars"] = len(str(meta.get("/Title") or os.path.basename(filepath)))
            f.seek(0)
            feats["valid_pdf_header"] = int(f.read(1024).decode(errors="ignore").startswith("%PDF"))
            f.seek(0)
            raw = f.read().decode("latin-1", errors="ignore")

        feats["endstream_count"] = raw.count("endstream")
        feats["stream_count"] = raw.count("stream")
        matches = list(re.finditer(r"stream(.*?)endstream", raw, re.DOTALL))
        sizes = [len(m.group(1)) for m in matches if m.group(1)]
        entropies = [_entropy(m.group(1)) for m in matches if m.group(1)]
        feats["average_stream_size"] = sum(sizes) / len(sizes) if sizes else 0
        feats["entropy_of_streams"] = sum(entropies) / len(entropies) if entropies else 0
        feats["name_obfuscations"] = _count_name_obfuscations(raw)

        kw_map = {
            "objstm_count": "/ObjStm", "js_count": "/JS", "javascript_count": "/JavaScript",
            "uri_count": "/URI", "action_count": "/Action", "aa_count": "/AA",
            "openaction_count": "/OpenAction", "launch_count": "/Launch",
            "submitform_count": "/SubmitForm", "acroform_count": "/AcroForm",
            "xfa_count": "/XFA", "jbig2decode_count": "/JBig2Decode",
            "colors_count": "/Colors", "richmedia_count": "/RichMedia",
            "trailer_count": "/Trailer", "xref_count": "/Xref",
            "startxref_count": "/startxref", "total_filters": "/Filter",
            "nested_filter_objects": "/Filter [",
        }
        for k, v in kw_map.items():
            feats[k] = raw.count(v)

        if re.search(r"https?://[^:\s]+:\d{4,5}", raw):
            feats["uses_nonstandard_port"] = 1

        behavior_kws = ["/JS", "/Launch", "/URI", "/OpenAction", "/SubmitForm", "/JavaScript", "/AA"]
        for block in re.findall(r"obj(.*?)endobj", raw, re.DOTALL):
            if sum(1 for bk in behavior_kws if bk in block) >= 2:
                feats["has_multiple_behavioral_keywords_in_one_object"] += 1

    except Exception as e:
        _log(log_path, filepath, f"PyPDF2 failed: {e}")

    # PyMuPDF stage
    try:
        import fitz
        doc = fitz.open(filepath)
        font_names: set[str] = set()
        for page in doc:
            feats["image_count"] += len(page.get_images(full=True))
            for fi in page.get_fonts():
                if fi[3]:
                    font_names.add(fi[3])
        feats["font_object_count"] = len(font_names)
        feats["object_count"] = doc.xref_length()
        feats["xref_entries"] = sum(
            1 for i in range(doc.xref_length()) if doc.xref_object(i, compressed=False)
        )
        feats["text_length"] = sum(len(p.get_text()) for p in doc)
        embedded: list[int] = []
        for i in range(doc.xref_length()):
            xobj = doc.xref_object(i, compressed=False)
            if "/EmbeddedFile" in xobj:
                try:
                    c = doc.xref_stream(i)
                    if c:
                        embedded.append(len(c))
                except Exception:
                    pass
        feats["embedded_file_count"] = len(embedded)
        feats["average_embedded_file_size"] = sum(embedded) / len(embedded) if embedded else 0
    except Exception as e:
        _log(log_path, filepath, f"PyMuPDF failed (fallback pdfminer): {e}")
        try:
            from pdfminer.high_level import extract_text
            text = extract_text(filepath) or ""
            feats["text_length"] = len(text)
            if not text.strip():
                feats["used_ocr"] = 1
        except Exception:
            pass

    return feats


# ─────────────────────────────────────────────────────────────────────────────
# Word extractor (43 features)
# ─────────────────────────────────────────────────────────────────────────────

WORD_FEATURE_ORDER = [
    "ole_object_count", "ole_object_type_count", "macro_present", "dde_present",
    "vba_keywords_count", "entropy", "struct_ContentType", "struct_PartName",
    "file_size", "struct_pos", "struct_val", "struct_typeface", "struct_script",
    "path_/w-p", "path_w-r", "path_/w-r", "path_a-hlink", "path_w-p",
    "path_a-accent3", "struct_ang", "path_/a-effectLst", "path_a-themeElements",
    "struct_dist", "path_a-alpha", "struct_Extension", "path_a-dk1", "path_a-ln",
    "path_/a-accent6", "struct_w", "struct_name", "path_a-lt2", "path_/a-outerShdw",
    "path_a-accent4", "path_/a-dk1", "path_a-accent1", "path_a-sysClr", "path_a-lt1",
    "path_/a-accent4",
    "struct_{http://schemas.openxmlformats.org/wordprocessingml/2006/main}sz",
    "path_a-solidFill",
    "struct_{http://schemas.openxmlformats.org/wordprocessingml/2006/main}themeFill",
    "struct_{http://schemas.openxmlformats.org/wordprocessingml/2006/main}csb1",
    "struct_{http://schemas.openxmlformats.org/wordprocessingml/2006/main}styleId",
]
WORD_COLS = ["file_path"] + WORD_FEATURE_ORDER + ["label"]

_VBA_KW_PATTERN = re.compile(
    r"\b(CreateObject|Shell|AppActivate|Environ|Execute|FileCopy|Dir|Kill|Put|Get|Open)\b",
    re.IGNORECASE,
)


def _count_ole_in_olefile(path: str) -> tuple[int, int]:
    try:
        import olefile
        if not olefile.isOleFile(path):
            return 0, 0
        with olefile.OleFileIO(path) as ole:
            names = ole.listdir(streams=True, storages=True)
            cands = ["/".join(n) for n in names if any(s.lower() in ("objectpool", "ole", "\x01ole") for s in n)]
            exts = {os.path.splitext(os.path.basename(n))[1].lower() for n in cands if os.path.splitext(os.path.basename(n))[1]}
            return len(cands), len(exts)
    except Exception:
        return 0, 0


def extract_word(filepath: str, label: int | None, log_path: Path) -> dict:
    feats = {col: 0 for col in WORD_COLS}
    feats["file_path"] = filepath
    feats["label"] = label if label is not None else ""
    try:
        feats["file_size"] = os.path.getsize(filepath)
    except OSError:
        pass

    text_all = ""

    try:
        have_olevba = False
        try:
            from oletools.olevba import VBA_Parser
            have_olevba = True
        except ImportError:
            pass

        if zipfile.is_zipfile(filepath):
            with zipfile.ZipFile(filepath, "r") as z:
                names = z.namelist()
                for name in names:
                    low = name.lower()
                    for key in WORD_FEATURE_ORDER:
                        if key.startswith("path_"):
                            needle = key[5:].lower()
                            if needle and needle in low:
                                feats[key] += 1

                xml_files = [n for n in names if n.lower().endswith(".xml")]
                for n in xml_files:
                    try:
                        with z.open(n) as xf:
                            xml_text = xf.read().decode(errors="ignore")
                    except Exception:
                        xml_text = ""
                    if not xml_text:
                        continue
                    text_all += xml_text
                    for key in WORD_FEATURE_ORDER:
                        if key.startswith("struct_"):
                            needle = key[7:]
                            if needle:
                                feats[key] += xml_text.count(needle)

                feats["entropy"] = _entropy(text_all)

                ole_like = [n for n in names if "/embeddings/" in n.lower()]
                feats["ole_object_count"] += len(ole_like)
                type_exts = {os.path.splitext(n)[1].lower() for n in ole_like if os.path.splitext(n)[1]}
                feats["ole_object_type_count"] += len(type_exts)

                if have_olevba:
                    vb = VBA_Parser(filepath)
                    try:
                        if vb.detect_vba_macros():
                            feats["macro_present"] = 1
                            vba_text = ""
                            for (_, _, _, code) in vb.extract_macros():
                                if code:
                                    vba_text += code + "\n"
                            if vba_text:
                                feats["vba_keywords_count"] = len(_VBA_KW_PATTERN.findall(vba_text))
                                feats["dde_present"] = int("DDEAUTO" in vba_text or "DDE" in vba_text)
                    finally:
                        vb.close()

        else:
            ext = os.path.splitext(filepath)[1].lower()
            if ext in (".doc", ".dot"):
                c, t = _count_ole_in_olefile(filepath)
                feats["ole_object_count"] += c
                feats["ole_object_type_count"] += t
                if have_olevba:
                    vb = VBA_Parser(filepath)
                    try:
                        if vb.detect_vba_macros():
                            feats["macro_present"] = 1
                            vba_text = ""
                            for (_, _, _, code) in vb.extract_macros():
                                if code:
                                    vba_text += code + "\n"
                            if vba_text:
                                feats["vba_keywords_count"] = len(_VBA_KW_PATTERN.findall(vba_text))
                                feats["dde_present"] = int("DDEAUTO" in vba_text or "DDE" in vba_text)
                                feats["entropy"] = _entropy(vba_text)
                    finally:
                        vb.close()

    except Exception as e:
        _log(log_path, filepath, f"Word extraction failed: {e}")

    return feats


# ─────────────────────────────────────────────────────────────────────────────
# Excel extractor (55 features)
# ─────────────────────────────────────────────────────────────────────────────

EXCEL_COLS = [
    "file_path", "file_size", "sheet_count", "max_rows", "max_cols",
    "total_cells", "non_empty_cells", "numeric_cell_count", "string_cell_count",
    "formula_count", "hyperlink_count", "avg_cell_length", "entropy_of_text",
    "base64_pattern_count", "hex_pattern_count", "has_macro",
    "remote_template_present", "ocr_extracted_text_length",
    "preview_image_text_entropy", "deceptive_keywords_count_ocr",
    "macro_line_count", "macro_procedure_count", "macro_chr_count",
    "macro_string_function_count", "macro_arithmetic_operator_count",
    "macro_concatenation_count", "macro_callbyname_count",
    "macro_comment_lines", "macro_average_line_length", "macro_token_count",
    "macro_count", "uses_file_api", "uses_network_api", "uses_process_api",
    "merged_cells_count", "hidden_sheets_count", "protected_sheets_count",
    "named_ranges_count", "empty_sheet_count", "rich_text_formatting_count",
    "macro_count_parentheses", "macro_count_assignments",
    "macro_max_line_length", "macro_max_string_literals",
    "macro_max_arithmetic_ops", "macro_max_concat_ops", "macro_vocab_size",
    "preview_image_width", "preview_image_height",
    "label",
]

_BASE64_RE = re.compile(r"(?:[A-Za-z0-9+/]{4}){10,}(?:[A-Za-z0-9+/]{2}==|[A-Za-z0-9+/]{3}=)?")
_HEX_RE    = re.compile(r"(?:\\x[0-9a-fA-F]{2}){4,}")
_PROC_RE   = re.compile(r"\b(Sub|Function)\b", re.IGNORECASE)
_SUSPICIOUS_OCR = ["enable content", "click here", "view document", "macro", "enable editing"]


def _is_openxml(path: str) -> bool:
    try:
        with zipfile.ZipFile(path) as z:
            return any(n.startswith("xl/") for n in z.namelist())
    except Exception:
        return False


def _extract_xlsx(filepath: str, feats: dict) -> list[str]:
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=False, keep_links=True)
    feats["sheet_count"] = len(wb.sheetnames)
    feats["named_ranges_count"] = len(list(wb.defined_names))
    all_text: list[str] = []
    for sheet in wb.worksheets:
        rows = sheet.max_row or 0
        cols = sheet.max_column or 0
        feats["max_rows"] = max(feats["max_rows"], rows)
        feats["max_cols"] = max(feats["max_cols"], cols)
        feats["total_cells"] += rows * cols
        if sheet.sheet_state != "visible":
            feats["hidden_sheets_count"] += 1
        if getattr(sheet, "protection", None) and sheet.protection.sheet:
            feats["protected_sheets_count"] += 1
        try:
            if not any(cell.value for row in sheet.iter_rows() for cell in row):
                feats["empty_sheet_count"] += 1
        except Exception:
            pass
        try:
            feats["merged_cells_count"] += len(sheet.merged_cells.ranges)
        except Exception:
            pass
        for row in sheet.iter_rows():
            for cell in row:
                val = str(cell.value).strip() if cell.value is not None else ""
                if val:
                    feats["non_empty_cells"] += 1
                    all_text.append(val)
                    if isinstance(cell.value, str):
                        feats["string_cell_count"] += 1
                    elif isinstance(cell.value, (int, float)):
                        feats["numeric_cell_count"] += 1
                    if cell.hyperlink:
                        feats["hyperlink_count"] += 1
                if cell.data_type == "f":
                    feats["formula_count"] += 1
                if getattr(cell, "font", None) and (cell.font.bold or cell.font.italic or cell.font.underline):
                    feats["rich_text_formatting_count"] += 1
    try:
        with zipfile.ZipFile(filepath) as z:
            for name in z.namelist():
                if "vbaProject.bin" in name:
                    feats["has_macro"] = 1
                    feats["macro_count"] += 1
                if name.lower().endswith(".xml"):
                    with z.open(name) as xf:
                        content = xf.read().decode(errors="ignore").lower()
                        if "http://" in content or "https://" in content:
                            feats["remote_template_present"] = 1
    except Exception:
        pass
    return all_text


def _extract_xls(filepath: str, feats: dict) -> list[str]:
    import xlrd
    wb = xlrd.open_workbook(filepath, formatting_info=False)
    feats["sheet_count"] = wb.nsheets
    all_text: list[str] = []
    for sheet in wb.sheets():
        rows, cols = sheet.nrows, sheet.ncols
        feats["max_rows"] = max(feats["max_rows"], rows)
        feats["max_cols"] = max(feats["max_cols"], cols)
        feats["total_cells"] += rows * cols
        for r in range(rows):
            for c in range(cols):
                try:
                    val = sheet.cell_value(r, c)
                except Exception:
                    val = ""
                if val not in ("", None):
                    feats["non_empty_cells"] += 1
                    sval = str(val).strip()
                    all_text.append(sval)
                    if isinstance(val, str):
                        feats["string_cell_count"] += 1
                    elif isinstance(val, (int, float)):
                        feats["numeric_cell_count"] += 1
    return all_text


def _extract_xlsb(filepath: str, feats: dict) -> list[str]:
    from pyxlsb import open_workbook
    all_text: list[str] = []
    sheet_count = 0
    with open_workbook(filepath) as wb:
        for sheet_name in wb.sheets:
            sheet_count += 1
            with wb.get_sheet(sheet_name) as sh:
                for row in sh.rows():
                    for cell in row:
                        v = cell.v
                        if v is not None and str(v).strip():
                            all_text.append(str(v).strip())
                            feats["non_empty_cells"] += 1
                            if isinstance(v, str):
                                feats["string_cell_count"] += 1
                            elif isinstance(v, (int, float)):
                                feats["numeric_cell_count"] += 1
    feats["sheet_count"] = sheet_count
    return all_text


def extract_excel(filepath: str, label: int | None, log_path: Path) -> dict:
    feats: dict = {col: 0 for col in EXCEL_COLS}
    feats["file_path"] = filepath
    feats["label"] = label if label is not None else ""
    feats["avg_cell_length"] = 0.0
    feats["entropy_of_text"] = 0.0
    feats["macro_average_line_length"] = 0.0
    feats["preview_image_text_entropy"] = 0.0
    try:
        feats["file_size"] = os.path.getsize(filepath)
    except OSError:
        pass

    ext = os.path.splitext(filepath)[1].lower()
    try:
        if ext in (".xlsx", ".xlsm") or _is_openxml(filepath):
            all_text = _extract_xlsx(filepath, feats)
        elif ext == ".xls":
            all_text = _extract_xls(filepath, feats)
        elif ext == ".xlsb":
            all_text = _extract_xlsb(filepath, feats)
        else:
            try:
                all_text = _extract_xlsx(filepath, feats)
            except Exception:
                all_text = _extract_xls(filepath, feats)

        joined = "\n".join(all_text)
        feats["avg_cell_length"] = (sum(len(t) for t in all_text) / len(all_text)) if all_text else 0.0
        feats["entropy_of_text"] = _entropy(joined)
        feats["base64_pattern_count"] = len(_BASE64_RE.findall(joined))
        feats["hex_pattern_count"] = len(_HEX_RE.findall(joined))

        lines = joined.splitlines()
        tokens = re.findall(r"\w+", joined)
        feats["macro_line_count"] = len(lines)
        feats["macro_procedure_count"] = sum(1 for l in lines if _PROC_RE.search(l))
        feats["macro_chr_count"] = joined.lower().count("chr")
        feats["macro_string_function_count"] = sum(
            joined.lower().count(f) for f in ["replace", "ucase", "lcase", "split", "instr", "strreverse"]
        )
        feats["macro_arithmetic_operator_count"] = sum(joined.count(op) for op in "+-*/")
        feats["macro_concatenation_count"] = joined.count("&")
        feats["macro_callbyname_count"] = joined.lower().count("callbyname")
        feats["macro_comment_lines"] = sum(1 for l in lines if l.strip().startswith("'"))
        feats["macro_average_line_length"] = (sum(len(l) for l in lines) / len(lines)) if lines else 0.0
        feats["macro_token_count"] = len(set(tokens))
        feats["macro_vocab_size"] = len(set(tokens))
        feats["macro_count_parentheses"] = joined.count("(") + joined.count(")")
        feats["macro_count_assignments"] = joined.count("=")
        feats["macro_max_line_length"] = max((len(l) for l in lines), default=0)
        feats["macro_max_string_literals"] = max((l.count('"') for l in lines), default=0)
        feats["macro_max_arithmetic_ops"] = max((sum(l.count(op) for op in "+-*/") for l in lines), default=0)
        feats["macro_max_concat_ops"] = max((l.count("&") for l in lines), default=0)
        feats["uses_file_api"] = int(any(k in joined for k in ["Open", "Write", "SaveAs", "FileCopy"]))
        feats["uses_network_api"] = int(any(k in joined for k in ["XMLHttpRequest", "WinHttpRequest", "URLDownloadToFile"]))
        feats["uses_process_api"] = int(any(k in joined for k in ["Shell", "CreateProcess", "Wscript.Shell"]))

        # OCR preview (best-effort, needs pillow + pytesseract)
        try:
            import pytesseract
            from PIL import Image
            with zipfile.ZipFile(filepath) as z:
                for name in z.namelist():
                    if name.lower().startswith("xl/media") and name.lower().endswith((".jpg", ".jpeg", ".png")):
                        with z.open(name) as img_file:
                            img = Image.open(img_file).convert("RGB")
                            ocr_text = pytesseract.image_to_string(img)
                            w, h = img.size
                            feats["ocr_extracted_text_length"] = len(ocr_text)
                            feats["preview_image_text_entropy"] = _entropy(ocr_text)
                            feats["preview_image_width"] = w
                            feats["preview_image_height"] = h
                            feats["deceptive_keywords_count_ocr"] = sum(
                                ocr_text.lower().count(p) for p in _SUSPICIOUS_OCR
                            )
                        break
        except Exception:
            pass

    except Exception as e:
        _log(log_path, filepath, f"Excel extraction failed: {e}")

    return feats


# ─────────────────────────────────────────────────────────────────────────────
# QR Code extractor (30 features)
# ─────────────────────────────────────────────────────────────────────────────

QR_COLS = [
    "file_path",
    # Image properties
    "file_size", "image_width", "image_height", "image_mode_id",
    "image_entropy", "brightness_mean", "contrast_std",
    "aspect_ratio",
    # QR decode
    "qr_detected", "qr_count", "decoded_text_length",
    "decoded_text_entropy",
    # URL analysis
    "url_present", "has_https", "has_http", "url_length",
    "url_entropy", "domain_length", "subdomain_count",
    "path_depth", "query_param_count", "has_fragment",
    "uses_ip_address", "uses_nonstandard_port",
    # URL suspicion indicators
    "special_char_count", "digit_ratio", "hyphen_count",
    "at_symbol_present", "double_slash_present",
    "label",
]

_IMAGE_MODE_MAP = {"1": 0, "L": 1, "P": 2, "RGB": 3, "RGBA": 4,
                   "CMYK": 5, "YCbCr": 6, "LAB": 7, "HSV": 8, "I": 9, "F": 10}
_IP_RE    = re.compile(r"^\d{1,3}(\.\d{1,3}){3}$")
_PORT_RE  = re.compile(r"https?://[^/\s]+:(\d+)")


def _entropy_bytes(data: bytes) -> float:
    if not data:
        return 0.0
    counts = Counter(data)
    total = len(data)
    return -sum((c / total) * math.log2(c / total) for c in counts.values())


def _decode_qr(img_path: str) -> list[str]:
    """Decode tất cả QR codes trong ảnh. Thử pyzbar trước, fallback opencv."""
    results: list[str] = []

    # Primary: pyzbar
    try:
        from PIL import Image
        from pyzbar.pyzbar import decode as pyzbar_decode
        img = Image.open(img_path).convert("RGB")
        for obj in pyzbar_decode(img):
            if obj.type == "QRCODE":
                results.append(obj.data.decode("utf-8", errors="ignore"))
        if results:
            return results
    except Exception:
        pass

    # Fallback: OpenCV
    try:
        import cv2
        img_cv = cv2.imread(img_path)
        if img_cv is not None:
            detector = cv2.QRCodeDetector()
            data, _, _ = detector.detectAndDecode(img_cv)
            if data:
                results.append(data)
    except Exception:
        pass

    return results


def _analyse_url(url: str) -> dict:
    """Trả về dict các URL features."""
    from urllib.parse import urlparse, parse_qs
    out = {
        "url_present": 1, "has_https": 0, "has_http": 0,
        "url_length": len(url), "url_entropy": _entropy(url),
        "domain_length": 0, "subdomain_count": 0,
        "path_depth": 0, "query_param_count": 0, "has_fragment": 0,
        "uses_ip_address": 0, "uses_nonstandard_port": 0,
        "special_char_count": len(re.findall(r"[^a-zA-Z0-9./:_\-?=&#%+]", url)),
        "digit_ratio": sum(c.isdigit() for c in url) / len(url) if url else 0.0,
        "hyphen_count": url.count("-"),
        "at_symbol_present": int("@" in url),
        "double_slash_present": int("//" in url[7:]),  # bỏ qua scheme://
    }
    try:
        parsed = urlparse(url)
        out["has_https"] = int(parsed.scheme == "https")
        out["has_http"]  = int(parsed.scheme == "http")
        host = parsed.hostname or ""
        out["domain_length"] = len(host)
        out["subdomain_count"] = max(0, host.count(".") - 1)
        out["uses_ip_address"] = int(bool(_IP_RE.match(host)))
        out["path_depth"] = len([p for p in parsed.path.split("/") if p])
        out["query_param_count"] = len(parse_qs(parsed.query))
        out["has_fragment"] = int(bool(parsed.fragment))
        port_match = _PORT_RE.search(url)
        if port_match:
            port = int(port_match.group(1))
            out["uses_nonstandard_port"] = int(port not in (80, 443, 8080, 8443))
    except Exception:
        pass
    return out


def extract_qr(filepath: str, label: int | None, log_path: Path) -> dict:
    """Trích xuất 30 features từ một ảnh QR Code."""
    feats: dict = {col: 0 for col in QR_COLS}
    feats["file_path"] = filepath
    feats["label"] = label if label is not None else ""
    feats["digit_ratio"] = 0.0
    feats["aspect_ratio"] = 0.0

    try:
        feats["file_size"] = os.path.getsize(filepath)
    except OSError as e:
        _log(log_path, filepath, f"Size check failed: {e}")
        return feats

    # Image properties
    try:
        from PIL import Image
        import numpy as np

        img = Image.open(filepath)
        w, h = img.size
        feats["image_width"]   = w
        feats["image_height"]  = h
        feats["aspect_ratio"]  = round(w / h, 4) if h else 0.0
        feats["image_mode_id"] = _IMAGE_MODE_MAP.get(img.mode, 0)

        # Entropy trên raw bytes
        with open(filepath, "rb") as f:
            feats["image_entropy"] = round(_entropy_bytes(f.read()), 4)

        # Brightness & contrast trên grayscale
        gray = np.array(img.convert("L"), dtype=float)
        feats["brightness_mean"] = round(float(gray.mean()), 4)
        feats["contrast_std"]    = round(float(gray.std()),  4)

    except Exception as e:
        _log(log_path, filepath, f"Image analysis failed: {e}")

    # QR decode
    try:
        decoded_list = _decode_qr(filepath)
        feats["qr_detected"] = int(len(decoded_list) > 0)
        feats["qr_count"]    = len(decoded_list)

        if decoded_list:
            text = decoded_list[0]  # phân tích QR đầu tiên
            feats["decoded_text_length"]  = len(text)
            feats["decoded_text_entropy"] = round(_entropy(text), 4)

            # URL analysis
            if re.match(r"https?://", text, re.IGNORECASE):
                url_feats = _analyse_url(text)
                feats.update(url_feats)
            elif re.match(r"[a-zA-Z][a-zA-Z0-9+\-.]*://", text):
                # non-http scheme (tel:, mailto:, etc.)
                feats["url_present"] = 1
                feats["url_length"]  = len(text)
                feats["url_entropy"] = round(_entropy(text), 4)

    except Exception as e:
        _log(log_path, filepath, f"QR decode failed: {e}")

    return feats


# ─────────────────────────────────────────────────────────────────────────────
# File routing
# ─────────────────────────────────────────────────────────────────────────────

def _file_type(path: Path) -> str | None:
    ext = path.suffix.lower()
    if ext in PDF_EXTS:
        return "pdf"
    if ext in WORD_EXTS:
        return "word"
    if ext in EXCEL_EXTS:
        return "excel"
    if ext in QR_EXTS:
        return "qr"
    return None


def _safe_tmp_name(archive_path: Path, member_name: str, idx: int) -> str:
    """Tạo tên file tạm an toàn - chỉ dùng cho trường hợp bắt buộc."""
    member_base = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(member_name).name)
    if not member_base:
        member_base = f"member_{idx:06d}"
    return f"{archive_path.stem}_{idx:06d}_{member_base}"


def _collect_from_zip(
    archive_path: Path,
    allowed: set[str],
    temp_root: Path,
    log_path: Path,
) -> list[tuple[Path, str, str]]:
    """
    Đọc các file phù hợp trong ZIP và trả về (temp_path, type, display_path).
    AN TOÀN: Chỉ giải nén file cần thiết, xóa ngay sau khi xử lý.
    """
    items: list[tuple[Path, str, str]] = []
    try:
        with zipfile.ZipFile(archive_path, "r") as zf:
            for idx, info in enumerate(zf.infolist()):
                if info.is_dir():
                    continue
                ftype = _file_type(Path(info.filename))
                if not ftype or ftype not in allowed:
                    continue

                # Chỉ giải nén khi cần xử lý (không lưu trữ lâu)
                tmp_name = _safe_tmp_name(archive_path, info.filename, idx)
                out_path = temp_root / tmp_name
                out_path.parent.mkdir(parents=True, exist_ok=True)
                
                # Giải nén với permission hạn chế (chỉ owner đọc/ghi)
                with zf.open(info) as src, open(out_path, "wb", opener=lambda path, flags: os.open(path, flags, 0o600)) as dst:
                    shutil.copyfileobj(src, dst)

                display_path = f"{archive_path}!/{info.filename}"
                items.append((out_path, ftype, display_path))
    except Exception as e:
        _log(log_path, str(archive_path), f"ZIP collect failed: {e}")
    return items


def _collect_from_tar(
    archive_path: Path,
    allowed: set[str],
    temp_root: Path,
    log_path: Path,
) -> list[tuple[Path, str, str]]:
    """
    Giải nén các file phù hợp trong TAR/TAR.GZ/TGZ/TAZ.
    AN TOÀN: Permission hạn chế (0o600), file tạm chỉ tồn tại trong thời gian xử lý.
    """
    items: list[tuple[Path, str, str]] = []
    try:
        with tarfile.open(archive_path, "r:*") as tf:
            for idx, member in enumerate(tf.getmembers()):
                if not member.isfile():
                    continue
                ftype = _file_type(Path(member.name))
                if not ftype or ftype not in allowed:
                    continue

                src = tf.extractfile(member)
                if src is None:
                    continue

                tmp_name = _safe_tmp_name(archive_path, member.name, idx)
                out_path = temp_root / tmp_name
                out_path.parent.mkdir(parents=True, exist_ok=True)
                
                # Giải nén với permission hạn chế (chỉ owner đọc/ghi)
                with src, open(out_path, "wb", opener=lambda path, flags: os.open(path, flags, 0o600)) as dst:
                    shutil.copyfileobj(src, dst)

                display_path = f"{archive_path}!/{member.name}"
                items.append((out_path, ftype, display_path))
    except Exception as e:
        _log(log_path, str(archive_path), f"TAR collect failed: {e}")
    return items


def _collect_from_gz(
    archive_path: Path,
    allowed: set[str],
    temp_root: Path,
    log_path: Path,
) -> list[tuple[Path, str, str]]:
    """
    Giải nén file .gz đơn.
    AN TOÀN: Permission hạn chế (0o600), file tạm chỉ tồn tại trong thời gian xử lý.
    """
    items: list[tuple[Path, str, str]] = []

    # .tar.gz/.tgz/.taz đã xử lý ở TAR path
    name_l = archive_path.name.lower()
    if name_l.endswith((".tar.gz", ".tgz", ".taz")):
        return items

    # Suy ra tên file bên trong từ tên archive
    member_name = archive_path.stem
    ftype = _file_type(Path(member_name))
    if not ftype or ftype not in allowed:
        return items

    try:
        tmp_name = _safe_tmp_name(archive_path, member_name, 0)
        out_path = temp_root / tmp_name
        out_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Giải nén với permission hạn chế (chỉ owner đọc/ghi)
        with gzip.open(archive_path, "rb") as src, open(out_path, "wb", opener=lambda path, flags: os.open(path, flags, 0o600)) as dst:
            shutil.copyfileobj(src, dst)
            
        display_path = f"{archive_path}!/{member_name}"
        items.append((out_path, ftype, display_path))
    except Exception as e:
        _log(log_path, str(archive_path), f"GZ collect failed: {e}")
    return items


def _collect(
    input_path: Path,
    allowed: set[str],
    temp_root: Path,
    log_path: Path,
) -> list[tuple[Path, str, str]]:
    """Chỉ đọc từ archive. Trả về list (local_path, type, display_path)."""
    files: list[tuple[Path, str, str]] = []

    def _is_archive(p: Path) -> bool:
        lname = p.name.lower()
        return (
            lname.endswith(".tar.gz")
            or lname.endswith(".tgz")
            or lname.endswith(".taz")
            or p.suffix.lower() in ARCHIVE_EXTS
        )

    def _add_archive(p: Path) -> None:
        print(f"  EXTRACT {p.name}...", end="           \r", flush=True)
        lname = p.name.lower()
        if zipfile.is_zipfile(p):
            try:
                found = _collect_from_zip(p, allowed, temp_root, log_path)
                print(f"  DONE {p.name}: {len(found)} file    ")
                files.extend(found)
            except Exception as e:
                print(f"  ERROR Zip {p.name}: {e}")
            return

        if tarfile.is_tarfile(p) or lname.endswith((".tar.gz", ".tgz", ".taz", ".tar")):
            try:
                found = _collect_from_tar(p, allowed, temp_root, log_path)
                print(f"  DONE {p.name}: {len(found)} file    ")
                files.extend(found)
            except Exception as e:
                print(f"  ERROR tar {p.name}: {e}")
            return
        
        if lname.endswith(".gz"):
            try:
                found = _collect_from_gz(p, allowed, temp_root, log_path)
                print(f"  DONE {p.name}: {len(found)} file    ")
                files.extend(found)
            except Exception as e:
                print(f"  ERROR GZ {p.name}: {e}")

    print(f"Đang quét tìm file nén trong: {input_path}...")
    if input_path.is_file():
        if _is_archive(input_path):
            _add_archive(input_path)
    else:
        for p in sorted(input_path.rglob("*")):
            if p.is_file() and _is_archive(p):
                _add_archive(p)
    print(f" DONE: Tìm thấy {len(files)} file phù hợp để xử lý.               ")
    return files


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    print("\n FEARTURE EXTRACTION")
    parser = argparse.ArgumentParser(
        description="Trích xuất features từ PDF / Word / Excel / QR Code → CSV",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--input",  "-i", required=True,
                        help="File nén hoặc thư mục chứa file nén (zip/tar/taz/gz)")
    parser.add_argument("--label",  "-l", type=int, default=None,
                        help="Nhãn (0=benign…4=critical). Bỏ trống = không gán nhãn")
    parser.add_argument("--types",  "-t", default="pdf,word,excel,qr",
                        help="Loại file cần xử lý, phân cách bằng dấu phẩy "
                             "(mặc định: pdf,word,excel,qr)")
    parser.add_argument("--output-dir", default=None,
                        help="Thư mục đầu ra cho CSV. Mặc định: dataset/Dataset/<type>/")
    parser.add_argument("--max-files", type=int, default=None,
                        help="Số file tối đa mỗi loại (không giới hạn mặc định)")
    parser.add_argument("--resume", action="store_true",
                        help="Bỏ qua file đã có trong CSV đầu ra")
    parser.add_argument("--error-log", default=None,
                        help="File log lỗi (mặc định: dataset/extract_errors.log)")
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    print(f"  [Input]  {input_path}")
    
    allowed    = {t.strip().lower() for t in args.types.split(",")}
    log_path   = Path(args.error_log) if args.error_log else SCRIPT_DIR / "extract_errors.log"
    temp_root  = SCRIPT_DIR / "_extract_tmp"

    valid_types = set(DEFAULT_OUT.keys())
    unknown_types = sorted(t for t in allowed if t and t not in valid_types)
    if unknown_types:
        print(f"ERROR: Type không hợp lệ: {', '.join(unknown_types)}")
        print(f"       Type hợp lệ: {', '.join(sorted(valid_types))}")
        sys.exit(1)
    allowed = {t for t in allowed if t in valid_types}

    if not input_path.exists():
        print(f"ERROR: Không tìm thấy: {input_path}")
        sys.exit(1)

    if temp_root.exists():
        shutil.rmtree(temp_root, ignore_errors=True)
    temp_root.mkdir(parents=True, exist_ok=True)

    # Build output paths
    out_paths: dict[str, Path] = {}
    for t in allowed:
        if args.output_dir:
            out_paths[t] = Path(args.output_dir) / f"{t}_features.csv"
        else:
            out_paths[t] = DEFAULT_OUT[t]
    for p in out_paths.values():
        p.parent.mkdir(parents=True, exist_ok=True)

    # Resume: đọc file_path đã xử lý
    processed: dict[str, set[str]] = {t: set() for t in allowed}
    if args.resume:
        for t, out_csv in out_paths.items():
            if out_csv.exists():
                try:
                    df_ex = pd.read_csv(out_csv, usecols=["file_path"])
                    processed[t] = set(df_ex["file_path"].astype(str).tolist())
                    print(f"  [Resume {t}] {len(processed[t])} file đã xử lý.")
                except Exception as e:
                    print(f"  [Cảnh báo {t}] Không đọc CSV cũ: {e}")

    # Collect
    print("\nGiải nén và trích xuất file từ archive...")
    all_files = _collect(input_path, allowed, temp_root, log_path)
    
    if args.max_files:
        print(f"  MAX: Lấy {args.max_files} file mỗi loại...")
        counts: dict[str, int] = {t: 0 for t in allowed}
        filtered = []
        for p, t, dp in all_files:
            if counts[t] < args.max_files:
                filtered.append((p, t, dp))
                counts[t] += 1
        all_files = filtered

    if not all_files:
        print("WARNING: Không tìm thấy file phù hợp bên trong archive (zip/tar/taz/gz).")
        try:
             shutil.rmtree(temp_root, ignore_errors=True)
        except: pass
        sys.exit(0)

    # Count per type
    type_count: dict[str, int] = Counter(t for _, t, _ in all_files)
    print("\n  Thống kê file tìm được:")
    for t, n in type_count.items():
        print(f"    - {t.upper():6s}: {n} file  →  {out_paths[t]}")
    print()

    # Extract
    print(f"Trích xuất đặc trưng")
    col_map  = {"pdf": PDF_COLS, "word": WORD_COLS, "excel": EXCEL_COLS, "qr": QR_COLS}
    ext_map  = {"pdf": extract_pdf, "word": extract_word, "excel": extract_excel, "qr": extract_qr}
    headers  = {t: not out_paths[t].exists() for t in allowed}
    success: dict[str, int] = {t: 0 for t in allowed}
    skipped: dict[str, int] = {t: 0 for t in allowed}
    total = len(all_files)

    # Open CSV handles
    csv_handles = {t: open(out_paths[t], "a", encoding="utf-8", newline="") for t in allowed}
    
    try:
        for idx, (local_path, ftype, display_path) in enumerate(all_files, 1):
            if display_path in processed[ftype]:
                skipped[ftype] += 1
                # print(f"  [{idx}/{total}] [Bỏ qua - Đã có] {display_path.split('!/')[-1]}")
                continue

            label_str = f"nhãn={args.label}" if args.label is not None else "không nhãn"
            source_name = display_path.split("!/")[-1]

            # Print status BEFORE processing
            print(f"  [{idx}/{total}] [{ftype.upper()}] Đang xử lý: {source_name}...", end=" ", flush=True)
            t0 = time.time()

            try:
                # Actual extraction call
                feats = ext_map[ftype](str(local_path), args.label, log_path)

                # Post-processing
                feats["file_path"] = display_path
                cols  = col_map[ftype]
                row   = pd.DataFrame([{c: feats.get(c, 0) for c in cols}])

                # Write to CSV immediately
                row.to_csv(csv_handles[ftype], index=False, header=headers[ftype])
                csv_handles[ftype].flush() # Ensure it's written to disk

                headers[ftype] = False
                success[ftype] += 1

                # Print success AFTER processing
                print(f"Xong ({time.time() - t0:.2f}s)")

            except Exception as e:
                _log(log_path, display_path, f"[{ftype}] {e}")
                print(f"Lỗi: {e}")

            finally:
                # XÓA FILE TẠM NGAY SAU KHI XỬ LÝ (dù thành công hay thất bại)
                if local_path.exists():
                    try:
                        os.remove(local_path)
                        print("Proccesing...", end="")
                    except Exception as cleanup_err:
                        print(f"Error: {cleanup_err}]", end="")
                print()  # newline

    except KeyboardInterrupt:
        print("\n\nStop...")
    finally:
        for fh in csv_handles.values():
            fh.close()
        # Clean up temp files
        print("Xóa thư mục tạm...")
        try:
            shutil.rmtree(temp_root, ignore_errors=True)
        except Exception as e:
            print(f" ERROR... {e}")

    print("\n Result sumary:")
    for t in allowed:
        print(f"  {t.upper():6s}: thành công={success[t]}  bỏ qua={skipped[t]}  → {out_paths[t]}")
    print(f"  Log lỗi: {log_path}")


if __name__ == "__main__":
    main()
